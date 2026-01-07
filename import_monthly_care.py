import sqlite3
from openpyxl import load_workbook
import re

DB_NAME = "baby.db"
EXCEL_FILE = "育兒系統_Database.xlsx"
SHEET_NAME = "提醒規則"

def extract_month(month_text: str):
    """
    從字串抓出月份數字：
    例如：寶寶滿八個月(抓不到) / 寶寶滿8個月(抓得到)
    目前用數字版最穩：1,2,3...12...
    """
    if not month_text:
        return None
    m = re.search(r"(\d+)", str(month_text))
    return int(m.group(1)) if m else None

def main():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # 確保表存在（你 db.py init_db 已經會建，這裡保險）
    c.execute("""
    CREATE TABLE IF NOT EXISTS monthly_care (
        month INTEGER PRIMARY KEY,
        content TEXT
    )
    """)

    wb = load_workbook(EXCEL_FILE)
    sheet = wb[SHEET_NAME]

    inserted = 0
    skipped = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        month_text = row[1]   # B 欄：寶寶滿X個月
        content = row[4]      # E 欄：提醒內容

        month = extract_month(month_text)

        if month is None or not content:
            skipped += 1
            continue

        # 同一個 month 如果有多列內容：合併在一起（用換行分隔）
        c.execute("SELECT content FROM monthly_care WHERE month=?", (month,))
        existing = c.fetchone()

        if existing and existing[0]:
            new_content = str(existing[0]).strip() + "\n\n" + str(content).strip()
        else:
            new_content = str(content).strip()

        c.execute("""
            INSERT OR REPLACE INTO monthly_care (month, content)
            VALUES (?, ?)
        """, (month, new_content))

        inserted += 1

    conn.commit()
    conn.close()

    print(f"✅ 匯入完成：寫入/更新 {inserted} 筆，略過 {skipped} 筆")
    print("你可以用：SELECT * FROM monthly_care ORDER BY month; 來檢查")

if __name__ == "__main__":
    main()